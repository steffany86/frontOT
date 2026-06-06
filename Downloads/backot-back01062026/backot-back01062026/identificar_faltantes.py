#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Identificar los 2 registros faltantes comparando BD vs Excel
"""

import sys
from openpyxl import load_workbook

# Cargar el Excel y obtener todas las órdenes de la Hoja 2
archivo = r"C:\Users\JOSUE CABRERA\Downloads\backot-back01062026\backot-back01062026\CARRENHO (2).xlsx"

print("=" * 80)
print("COMPARACIÓN BD vs EXCEL - Buscando registros faltantes")
print("=" * 80)

try:
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Hoja2']
    
    # Obtener todos los números de orden del Excel (columna 3 - NRO_ORDEN)
    ordenes_excel = set()
    for row_idx in range(2, ws.max_row + 1):
        nro_orden = ws.cell(row_idx, 3).value
        if nro_orden:
            ordenes_excel.add(str(nro_orden))
    
    print(f"\nTotal de órdenes en Excel (Hoja 2): {len(ordenes_excel)}")
    print(f"\nPrimeras 10 órdenes en Excel:")
    for idx, orden in enumerate(sorted(list(ordenes_excel))[:10], 1):
        print(f"  {idx}. {orden}")
    
    wb.close()
    
    # Ahora consultar la BD para obtener las 40 órdenes
    print(f"\n{'=' * 80}")
    print("Consultando las 40 órdenes de la BD...")
    print(f"{'=' * 80}")
    
    # Generar script SQL para obtener las órdenes
    sql_script = """
USE BDControlOrdenes;
GO

DECLARE @F1 DATE = '2026-05-01', @F2 DATE = '2026-05-31';

WITH base AS (
    SELECT 
        r.nro_orden,
        r.fecha_de_respuesta,
        r.cliente_nombre_completo,
        r.tecnico_nombre,
        ROW_NUMBER() OVER (PARTITION BY r.id_transaccion ORDER BY CONVERT(DATETIME, r.fecha_carga, 103) DESC, r.id_NPS_RESPUESTAS_MAKIRO DESC) rn
    FROM dbo.tbl_NPS_RESPUESTAS_MAKIRO r
    WHERE CONVERT(DATE, r.fecha_de_respuesta, 103) BETWEEN @F1 AND @F2
      AND r.tecnico_nombre LIKE '%DANIEL CARREÑO ROJAS%'
)
SELECT 
    nro_orden,
    CONVERT(VARCHAR(10), fecha_de_respuesta, 103) AS fecha_respuesta,
    cliente_nombre_completo
FROM base 
WHERE rn=1
ORDER BY nro_orden;
"""
    
    with open('obtener_ordenes_bd.sql', 'w', encoding='utf-8') as f:
        f.write(sql_script)
    
    print("\nScript SQL generado: obtener_ordenes_bd.sql")
    print("Ejecuta el siguiente comando para obtener las órdenes de la BD:")
    print('sqlcmd -S 172.16.0.13 -d BDControlOrdenes -U sistemas -P sametsis -i obtener_ordenes_bd.sql -W -o ordenes_bd.txt')
    
    print(f"\nOrdenes en Excel guardadas en archivo para comparación posterior.")
    
    # Guardar las órdenes del Excel en un archivo
    with open('ordenes_excel.txt', 'w', encoding='utf-8') as f:
        for orden in sorted(ordenes_excel):
            f.write(f"{orden}\n")
    
    print("Órdenes del Excel guardadas en: ordenes_excel.txt")
    
except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
