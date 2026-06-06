import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('CARRENHO (2).xlsx')
ws = wb['Hoja2']

print('Órdenes que NO están en el Excel (29160908, 29162712):')
print('Revisando si alguna orden tiene características especiales...\n')

# Buscar columna de fecha_de_respuesta
fecha_resp_col = None
for j in range(1, ws.max_column + 1):
    if ws.cell(1, j).value and 'FECHA_DE_RESPUESTA' in str(ws.cell(1, j).value).upper():
        fecha_resp_col = j
        break

if fecha_resp_col:
    fechas = {}
    for i in range(2, ws.max_row + 1):
        nro_orden = ws.cell(i, 3).value  
        fecha_resp = ws.cell(i, fecha_resp_col).value
        if nro_orden:
            fechas[str(nro_orden)] = str(fecha_resp)
    
    # Ver fechas únicas
    from collections import Counter
    print('Distribución de fechas_de_respuesta en Excel:')
    counter = Counter(fechas.values())
    for fecha, count in sorted(counter.items()):
        print(f'{fecha}: {count} órdenes')
    
    # Ver específicamente si hay un patrón en las fechas
    print('\n\nFechas de respuesta tempranas (primeros 10 días de mayo):')
    tempranas = [orden for orden, fecha in fechas.items() if '05/2026' in fecha or '10/05/2026' in fecha]
    print(f'Total: {len(tempranas)} órdenes')
    if len(tempranas) <= 20:
        for orden in sorted(tempranas):
            print(f'  {orden}: {fechas[orden]}')
