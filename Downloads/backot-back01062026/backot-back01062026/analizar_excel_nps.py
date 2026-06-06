#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analizar el archivo Excel de Daniel Carreño para comparar con los datos de la BD
"""

import sys
import os

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no está instalado")
    print("Instala con: pip install openpyxl")
    sys.exit(1)

def analizar_excel(archivo_excel):
    """Analiza el archivo Excel y muestra información de todas las hojas"""
    
    if not os.path.exists(archivo_excel):
        print(f"ERROR: No se encuentra el archivo {archivo_excel}")
        return
    
    print("=" * 80)
    print(f"ANÁLISIS DE: {archivo_excel}")
    print("=" * 80)
    
    try:
        wb = load_workbook(archivo_excel, data_only=True)
        
        print(f"\nNúmero de hojas: {len(wb.sheetnames)}")
        print(f"Nombres de hojas: {wb.sheetnames}")
        print()
        
        # Analizar cada hoja
        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print("=" * 80)
            print(f"HOJA {idx}: {sheet_name}")
            print("=" * 80)
            
            # Contar filas con datos
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Contar filas no vacías
            filas_con_datos = 0
            for row in ws.iter_rows(min_row=1, max_row=max_row):
                if any(cell.value is not None for cell in row):
                    filas_con_datos += 1
            
            print(f"Dimensiones: {filas_con_datos} filas con datos x {max_col} columnas")
            print(f"Rango: A1:{ws.cell(max_row, max_col).coordinate}")
            
            # Mostrar encabezados (primera fila)
            print("\nEncabezados (primera fila):")
            headers = []
            for col in range(1, min(max_col + 1, 20)):  # Máximo 20 columnas
                cell_value = ws.cell(1, col).value
                if cell_value:
                    headers.append(str(cell_value)[:30])
                    print(f"  Col {col}: {str(cell_value)[:50]}")
            
            # Mostrar primeras 5 filas de datos
            print(f"\nPrimeras 5 filas de datos:")
            for row_idx in range(2, min(7, max_row + 1)):
                print(f"\n  Fila {row_idx}:")
                for col_idx in range(1, min(max_col + 1, 10)):
                    cell_value = ws.cell(row_idx, col_idx).value
                    if cell_value:
                        print(f"    Col {col_idx}: {str(cell_value)[:50]}")
            
            print()
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Buscar el archivo Excel
    directorio = os.path.dirname(os.path.abspath(__file__))
    
    # Buscar archivos que coincidan
    archivos_posibles = [
        os.path.join(directorio, "CARRENHO (2).xlsx"),
        os.path.join(directorio, "CARREÑO (2).xlsx"),
        os.path.join(os.path.dirname(directorio), "CARRENHO (2).xlsx"),
        os.path.join(os.path.dirname(directorio), "CARREÑO (2).xlsx"),
    ]
    
    archivo_encontrado = None
    for archivo in archivos_posibles:
        if os.path.exists(archivo):
            archivo_encontrado = archivo
            break
    
    if archivo_encontrado:
        analizar_excel(archivo_encontrado)
    else:
        print("No se encontró el archivo Excel.")
        print("Archivos buscados:")
        for archivo in archivos_posibles:
            print(f"  - {archivo}")
        print("\nColoca el archivo 'CARRENHO (2).xlsx' en el directorio del proyecto.")
