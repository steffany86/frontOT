import openpyxl
import sys

# Cargar el Excel
wb = openpyxl.load_workbook(r'C:\Users\JOSUE CABRERA\Downloads\CARRENHO (2).xlsx', data_only=True)
ws = wb['Hoja2']

print(f"Total de filas en Hoja2: {ws.max_row}")
print(f"Total de columnas: {ws.max_column}")
print()

# Extraer los encabezados
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    headers.append(header if header else f"Col_{col}")

print("=== ENCABEZADOS (primeros 20) ===")
for i, h in enumerate(headers[:20], 1):
    print(f"{i}. {h}")
print()

# Extraer todas las órdenes
ordenes_excel = []
for row in range(2, ws.max_row + 1):
    nro_orden = str(ws.cell(row=row, column=3).value)  # NRO_ORDEN está en columna 3
    if nro_orden and nro_orden != 'None':
        ordenes_excel.append(nro_orden)

print(f"Total de órdenes en Excel: {len(ordenes_excel)}")
print()

# Las 2 órdenes que están en BD pero NO en Excel
ordenes_faltantes = ['29160908', '29162712']
ordenes_presentes = ['29119172', '29122286']  # Para comparar

print("=== BUSCANDO SI LAS ÓRDENES FALTANTES ESTÁN EN EL EXCEL ===")
for orden in ordenes_faltantes:
    if orden in ordenes_excel:
        print(f"✓ Orden {orden} SÍ está en el Excel")
    else:
        print(f"✗ Orden {orden} NO está en el Excel")

print()
print("=== ÓRDENES DE CONTROL (deben estar) ===")
for orden in ordenes_presentes:
    if orden in ordenes_excel:
        print(f"✓ Orden {orden} SÍ está en el Excel")
    else:
        print(f"✗ Orden {orden} NO está en el Excel")

print()
print("=== VERIFICANDO SI HAY FILTROS U OCULTOS ===")
if ws.auto_filter:
    print(f"Auto-filter detectado: {ws.auto_filter.ref}")
else:
    print("No hay auto-filter detectado")

# Ver si hay filas ocultas
hidden_rows = []
for row_idx in range(2, min(50, ws.max_row + 1)):
    row_dim = ws.row_dimensions[row_idx]
    if row_dim.hidden:
        hidden_rows.append(row_idx)

if hidden_rows:
    print(f"Filas ocultas encontradas: {hidden_rows}")
else:
    print("No hay filas ocultas (revisadas primeras 48)")

print()
print("=== PRIMERAS 5 ÓRDENES DEL EXCEL ===")
for row in range(2, min(7, ws.max_row + 1)):
    nro_orden = ws.cell(row=row, column=3).value
    fecha_creacion = ws.cell(row=row, column=1).value
    fecha_respuesta = ws.cell(row=row, column=4).value
    print(f"Fila {row}: Orden={nro_orden}, Fecha_Creacion={fecha_creacion}, Fecha_Respuesta={fecha_respuesta}")
