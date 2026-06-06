import openpyxl

# Cargar el Excel
wb = openpyxl.load_workbook(r'C:\Users\JOSUE CABRERA\Downloads\CARRENHO (2).xlsx', data_only=True)
ws = wb['Hoja2']

# Extraer encabezados
headers = []
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value
    headers.append(header if header else f"Col_{col}")

# Buscar órdenes específicas
target_orders = ['29119172', '29160908', '29162712']
print("=== BUSCANDO ÓRDENES ESPECÍFICAS EN EL EXCEL ===\n")

for row in range(2, ws.max_row + 1):
    nro_orden = str(ws.cell(row=row, column=3).value)
    
    if nro_orden in target_orders:
        print(f"=== ORDEN {nro_orden} (Fila {row}) ===")
        for col_idx, header in enumerate(headers[:20], 1):  # Primeros 20 campos
            value = ws.cell(row=row, column=col_idx).value
            print(f"  {header}: {value}")
        print()

# Listar todas las órdenes del Excel ordenadas
all_orders = []
for row in range(2, ws.max_row + 1):
    nro_orden = str(ws.cell(row=row, column=3).value)
    fecha_creacion = ws.cell(row=row, column=1).value
    if nro_orden and nro_orden != 'None':
        all_orders.append((nro_orden, fecha_creacion))

all_orders.sort()
print("\n=== TODAS LAS ÓRDENES EN EL EXCEL (ordenadas) ===")
for orden, fecha in all_orders:
    print(f"{orden} - {fecha}")
