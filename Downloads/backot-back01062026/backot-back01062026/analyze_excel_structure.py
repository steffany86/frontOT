import openpyxl

wb = openpyxl.load_workbook('CARRENHO (2).xlsx')
ws = wb['Hoja2']

print('Columnas del Excel:')
headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
for i, h in enumerate(headers, 1):
    print(f'{i}. {h}')

# Ver algunos registros de fecha_carga
print('\n\nPrimeros 5 registros con nro_orden y fecha_carga:')
for i in range(2, min(7, ws.max_row + 1)):
    nro_orden = ws.cell(i, 3).value  # Columna 3 es nro_orden
    # Buscar columna fecha_carga
    fecha_carga_col = None
    for j, h in enumerate(headers, 1):
        if h and 'fecha_carga' in str(h).lower():
            fecha_carga_col = j
            break
    if fecha_carga_col:
        fecha_carga = ws.cell(i, fecha_carga_col).value
        print(f'Orden {nro_orden}: fecha_carga = {fecha_carga}')
