import openpyxl

wb = openpyxl.load_workbook('CARRENHO (2).xlsx')
ws = wb['Hoja2']

# Buscar columna de fecha_de_respuesta
fecha_resp_col = None
for j in range(1, ws.max_column + 1):
    if ws.cell(1, j).value and 'FECHA_DE_RESPUESTA' in str(ws.cell(1, j).value).upper():
        fecha_resp_col = j
        break

print('Órdenes con fecha_de_respuesta = 10/05/2026 en el Excel:')
for i in range(2, ws.max_row + 1):
    nro_orden = ws.cell(i, 3).value  
    fecha_resp = ws.cell(i, fecha_resp_col).value
    if fecha_resp and '2026-05-10' in str(fecha_resp):
        print(f'Orden: {nro_orden}, Fecha: {fecha_resp}')
