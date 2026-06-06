import openpyxl

wb = openpyxl.load_workbook('CARRENHO (2).xlsx')
ws = wb['Hoja2']

ordenes = []
for i in range(2, ws.max_row + 1):
    if ws.cell(i, 3).value:
        ordenes.append(str(ws.cell(i, 3).value))

ordenes_sorted = sorted(ordenes)

with open('ordenes_excel.txt', 'w') as f:
    f.write('\n'.join(ordenes_sorted))

print('\n'.join(ordenes_sorted))
print(f'\n\nTotal: {len(ordenes)} ordenes')
