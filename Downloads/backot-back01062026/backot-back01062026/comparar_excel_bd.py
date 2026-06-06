#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Comparar los datos del Excel con los de la base de datos
"""

import sys
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl no está instalado")
    print("Instala con: pip install openpyxl")
    sys.exit(1)

def analizar_hoja2(archivo_excel):
    """Analiza la Hoja 2 del Excel con detalle"""
    
    print("=" * 80)
    print("ANÁLISIS DETALLADO - HOJA 2")
    print("=" * 80)
    
    try:
        wb = load_workbook(archivo_excel, data_only=True)
        ws = wb['Hoja2']
        
        # Obtener encabezados
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header:
                headers.append((col, str(header).strip()))
        
        print(f"\nTotal de columnas con encabezado: {len(headers)}")
        print("\nEncabezados principales:")
        for col_idx, header in headers[:20]:
            print(f"  {col_idx}. {header}")
        
        # Contar filas con datos (excluyendo encabezado)
        registros = []
        mayo_2026 = 0
        otros_meses = 0
        
        for row_idx in range(2, ws.max_row + 1):
            # Verificar si la fila tiene datos
            fecha_creacion = ws.cell(row_idx, 1).value
            id_transaccion = ws.cell(row_idx, 2).value
            nro_orden = ws.cell(row_idx, 3).value
            fecha_respuesta = ws.cell(row_idx, 5).value
            nombre_cliente = ws.cell(row_idx, 8).value
            
            if id_transaccion or nro_orden:
                registro = {
                    'fila': row_idx,
                    'fecha_creacion': fecha_creacion,
                    'id_transaccion': id_transaccion,
                    'nro_orden': nro_orden,
                    'fecha_respuesta': fecha_respuesta,
                    'nombre_cliente': nombre_cliente
                }
                registros.append(registro)
                
                # Verificar si es de mayo 2026
                if fecha_respuesta:
                    try:
                        if isinstance(fecha_respuesta, datetime):
                            fecha_dt = fecha_respuesta
                        else:
                            fecha_str = str(fecha_respuesta)
                            if '2026-05-' in fecha_str or '/05/2026' in fecha_str or '-05-2026' in fecha_str:
                                mayo_2026 += 1
                                continue
                            fecha_dt = datetime.strptime(fecha_str.split()[0], '%Y-%m-%d')
                        
                        if fecha_dt.year == 2026 and fecha_dt.month == 5:
                            mayo_2026 += 1
                        else:
                            otros_meses += 1
                    except:
                        pass
        
        print(f"\n{'=' * 80}")
        print("RESUMEN DE REGISTROS")
        print(f"{'=' * 80}")
        print(f"Total de registros en Hoja2: {len(registros)}")
        print(f"Registros de Mayo 2026: {mayo_2026}")
        print(f"Registros de otros meses: {otros_meses}")
        
        print(f"\n{'=' * 80}")
        print("PRIMEROS 10 REGISTROS DE MAYO 2026")
        print(f"{'=' * 80}")
        
        mayo_count = 0
        for reg in registros:
            fecha_resp = reg['fecha_respuesta']
            if fecha_resp:
                try:
                    if isinstance(fecha_resp, datetime):
                        fecha_dt = fecha_resp
                    else:
                        fecha_str = str(fecha_resp)
                        if '2026-05-' in fecha_str or '/05/2026' in fecha_str:
                            fecha_dt = None
                        else:
                            fecha_dt = datetime.strptime(fecha_str.split()[0], '%Y-%m-%d')
                    
                    if fecha_dt and fecha_dt.year == 2026 and fecha_dt.month == 5:
                        mayo_count += 1
                        if mayo_count <= 10:
                            print(f"\n{mayo_count}. Fila {reg['fila']}")
                            print(f"   Orden: {reg['nro_orden']}")
                            print(f"   Cliente: {reg['nombre_cliente']}")
                            print(f"   Fecha respuesta: {fecha_resp}")
                    elif '2026-05-' in str(fecha_resp) or '/05/2026' in str(fecha_resp):
                        mayo_count += 1
                        if mayo_count <= 10:
                            print(f"\n{mayo_count}. Fila {reg['fila']}")
                            print(f"   Orden: {reg['nro_orden']}")
                            print(f"   Cliente: {reg['nombre_cliente']}")
                            print(f"   Fecha respuesta: {fecha_resp}")
                except:
                    pass
        
        print(f"\n{'=' * 80}")
        print("COMPARACIÓN CON BASE DE DATOS")
        print(f"{'=' * 80}")
        print(f"Registros en Excel (Hoja 2): {len(registros)}")
        print(f"Registros de Mayo 2026 en Excel: {mayo_2026}")
        print(f"Registros de Mayo 2026 en BD: 40")
        print(f"Diferencia: {40 - mayo_2026} registro(s)")
        
        if 40 - mayo_2026 == 1:
            print("\n¡Diferencia de 1 registro! Posibles razones:")
            print("  - Un registro duplicado que en la BD se elimina con ROW_NUMBER()")
            print("  - Un registro con fecha en el límite (31/05/2026 vs 01/06/2026)")
            print("  - Diferencia en zona horaria o formato de fecha")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    archivo = r"C:\Users\JOSUE CABRERA\Downloads\backot-back01062026\backot-back01062026\CARRENHO (2).xlsx"
    analizar_hoja2(archivo)
